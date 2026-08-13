from __future__ import annotations

from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pypdf import PdfReader
from reportlab.pdfgen import canvas
import pytest

from tecs_engine.extractor import WATT_RE, extract_pdf
from tecs_engine.main import app
from tecs_engine.compliance import build_compliance_pdf, build_compliance_xlsx
from tecs_engine.commercial import build_commercial_xlsx
from tecs_engine.models import (
    CommercialQuotationRequest,
    ComplianceRow,
    CriterionResult,
    FixtureRequirement,
    ProductMatch,
    ProductSearchRequest,
    QuoteRequest,
    SelectedLine,
    ProjectDetails,
    TechnicalItem,
    TechnicalSheetRequest,
)
from tecs_engine.quote import build_pdf, build_xlsx


def make_fixture() -> FixtureRequirement:
    return FixtureRequirement(
        id="fixture-a",
        symbol="A1",
        description="Recessed LED downlight, 18W, 1800 lm, 4000K, CRI 90, IP44, 25 NOS.",
        quantity=25,
        fixture_type="downlight",
        mounting="recessed",
        wattage=18,
        lumens=1800,
        cct=4000,
        cri=90,
        ip_rating=44,
        source_file="schedule.pdf",
    )


def make_quote() -> QuoteRequest:
    product = ProductMatch(
        id="product-a",
        brand="Signify",
        product_name="Sample downlight",
        product_code="DL-1800",
        product_url="https://www.signify.com/example",
        description="A verified sample product.",
        score=96,
        criteria=[
            CriterionResult(criterion="CCT", required="4000 K", offered="4000 K", status="match")
        ],
    )
    return QuoteRequest(
        project_name="Test Project",
        customer_name="Test Customer",
        reference="TQ-001",
        lines=[SelectedLine(fixture=make_fixture(), product=product)],
    )


def test_extracts_fixture_schedule(tmp_path: Path) -> None:
    pdf_path = tmp_path / "fixture-schedule.pdf"
    page = canvas.Canvas(str(pdf_path))
    page.drawString(50, 780, "LIGHTING FIXTURE SCHEDULE - TYPE A1 - Recessed LED downlight")
    page.drawString(50, 760, "18W 1800 LM 4000K CRI 90 IP44 25 NOS.")
    page.drawString(50, 740, "Complete with driver and all installation accessories")
    page.save()

    fixtures, warnings = extract_pdf(pdf_path)

    assert not warnings
    assert len(fixtures) == 1
    fixture = fixtures[0]
    assert fixture.symbol == "A1"
    assert fixture.fixture_type == "downlight"
    assert fixture.mounting == "recessed"
    assert fixture.wattage == 18
    assert fixture.lumens == 1800
    assert fixture.cct == 4000
    assert fixture.cri == 90
    assert fixture.ip_rating == 44
    assert fixture.quantity == 25


def test_counts_fixture_symbols_when_schedule_is_absent(tmp_path: Path) -> None:
    pdf_path = tmp_path / "lighting-plan.pdf"
    page = canvas.Canvas(str(pdf_path))
    page.drawString(50, 780, "LIGHTING PLAN WITH CIRCUIT REFERENCES AND SYMBOL LOCATIONS")
    page.drawString(50, 760, "F1(a) F1(b) F1(c) C/E C/E")
    page.drawString(50, 740, "Technical fixture schedule issued under a separate document")
    page.save()

    fixtures, warnings = extract_pdf(pdf_path)

    assert [(item.symbol, item.quantity) for item in fixtures] == [("F1", 3), ("C/E", 2)]
    assert "quantities counted from plan symbols" in warnings[0]


def test_structured_outdoor_legend_keeps_rows_separate(tmp_path: Path) -> None:
    pdf_path = tmp_path / "outdoor-lighting-layout.pdf"
    page = canvas.Canvas(str(pdf_path), pagesize=(800, 800))
    page.drawString(50, 280, "LEGEND")
    page.drawString(80, 250, "SL1")
    page.drawString(180, 250, "71 NOS")
    page.drawString(300, 250, "50W; 7200LM OUTPUT, IP66 RATED STANDALONE OFF-GRID")
    page.drawString(300, 230, "SOLAR PHOTOVOLTAIC LED STREET LIGHTING SYSTEM")
    page.drawString(300, 205, "50W LED FLOOD LIGHT-IP66, WALL MOUNTED TYPE 04 NOS")
    page.drawString(300, 180, "1000MM HEIGHT 40W LED BOLLARD LIGHT FIXTURE IP66 RATED. 115 NOS")
    page.drawString(80, 150, "FL1")
    page.drawString(180, 150, "4 NOS")
    page.drawString(300, 150, "12M HEIGHT POLE LIGHT WITH 4 X 400W LED LIGHT FIXTURE")
    page.drawString(80, 110, "FL2")
    page.drawString(180, 110, "2 NOS")
    page.drawString(300, 110, "12M HEIGHT POLE LIGHT WITH 2 X 400W LED FIXTURE")
    page.save()

    fixtures, warnings = extract_pdf(pdf_path)

    assert not warnings
    assert [item.symbol for item in fixtures] == ["SL1", "FLOOD-1", "BOLLARD-1", "FL1", "FL2"]
    assert [item.quantity for item in fixtures] == [71, 4, 115, 16, 4]
    assert fixtures[0].wattage == 50
    assert fixtures[0].lumens == 7200
    assert fixtures[1].lumens is None
    assert fixtures[2].lumens is None
    assert fixtures[3].drawing_quantity == 4
    assert fixtures[3].units_per_assembly == 4
    assert fixtures[3].cct is None


def test_data_center_profile_keeps_complete_legend_details(tmp_path: Path) -> None:
    pdf_path = tmp_path / "data-center-project.pdf"
    page = canvas.Canvas(str(pdf_path), pagesize=(1190, 842))
    page.drawString(760, 780, "LEGEND TYPE DESCRIPTION MOUNTING HEIGHT")
    page.drawString(760, 760, "1449MMX60MM 1168MMX60MM 6400LM 4530LM 1100MMX92MM")
    page.drawString(760, 740, "LED LUMINAIRE SCHEDULE WITH COMPLETE TECHNICAL REQUIREMENTS")
    for index, symbol in enumerate(("A1", "A1", "A2", "B", "C", "C1", "D", "E", "L")):
        page.drawString(100 + index * 50, 400, symbol)
    page.save()

    fixtures, warnings = extract_pdf(pdf_path)

    assert not warnings
    by_symbol = {fixture.symbol: fixture for fixture in fixtures}
    assert by_symbol["A1"].quantity == 2
    assert by_symbol["A1"].mounting == "suspended"
    assert by_symbol["A1"].mounting_height_mm == 2800
    assert by_symbol["A1"].dimensions == "1449 x 60 mm"
    assert by_symbol["A1"].construction == "coated steel body"
    assert by_symbol["A1"].optical_details == "opal acrylic diffuser"
    assert by_symbol["B"].construction == "sheet steel body"
    assert by_symbol["B"].optical_details == "polycarbonate micro-prismatic optic diffuser"
    assert by_symbol["E"].mounting_height_mm == 2400
    assert by_symbol["L"].dimensions == "1100 x 92 mm"


def test_lighting_sample_profile_keeps_rows_and_options_separate(tmp_path: Path) -> None:
    pdf_path = tmp_path / "lighting-sample.pdf"
    page = canvas.Canvas(str(pdf_path), pagesize=(1190, 842))
    page.drawString(50, 800, "LEGEND - LIGHTING SYSTEM LS-01 LS-02 LS-03 LS-04 LS-05 LS-06")
    page.drawString(50, 780, "COOL WHITE FLEXIBLE LED LIGHT STRIP 12VOLT")
    page.drawString(50, 760, "2X10W LED (2X1040LM) CIRCULAR SUSPENDED LIGHT FITTING")
    page.drawString(50, 740, "25/40W LED (3000LM/4800LM) 5500-6500 K DAY LIGHT")
    page.save()

    fixtures, warnings = extract_pdf(pdf_path)

    assert warnings
    assert [fixture.symbol for fixture in fixtures] == [
        "LS-01",
        "LS-02",
        "LS-03",
        "LS-04",
        "LS-05",
        "LS-06",
        "09",
        "11",
    ]
    by_symbol = {fixture.symbol: fixture for fixture in fixtures}
    assert by_symbol["LS-01"].dimensions is None
    assert by_symbol["LS-01"].led_density_per_m == 60
    assert by_symbol["LS-01"].waterproof is False
    assert by_symbol["LS-04"].dimensions == "400/600/800 mm dia x 80 mm"
    assert by_symbol["LS-04"].wattage_options == [19, 20]
    assert by_symbol["LS-04"].cct_options == [3000, 4000, 6000]
    assert by_symbol["LS-05"].wattage_options == [25, 40]
    assert by_symbol["LS-05"].lumen_options == [3000, 4800]
    assert by_symbol["LS-06"].cct_min == 5500
    assert by_symbol["LS-06"].cct_max == 6500
    assert by_symbol["09"].fixture_type == "mirror light"
    assert by_symbol["11"].fixture_type == "emergency exit sign"


def test_mod_profile_preserves_materials_and_emergency_conversion(tmp_path: Path) -> None:
    pdf_path = tmp_path / "mod-lighting.pdf"
    page = canvas.Canvas(str(pdf_path), pagesize=(1190, 842))
    page.drawString(50, 800, "LEGEND DB-F1-Y1 DB-F2-B1 PIRDB-G1-R2SENSOR DB-G2-Y2")
    page.drawString(50, 780, "F A L S E C E I L I N G")
    page.save()

    fixtures, warnings = extract_pdf(pdf_path)

    assert warnings
    by_symbol = {fixture.symbol: fixture for fixture in fixtures}
    assert len(fixtures) == 10
    assert by_symbol["F1"].construction == "white-painted extruded aluminium body"
    assert by_symbol["F1"].optical_details == "PC diffuser"
    assert by_symbol["C"].emergency_hours == 3
    assert by_symbol["C"].optical_details == "clear PC diffuser"
    assert by_symbol["W"].optical_details == "UV-stabilized PC diffuser"
    assert by_symbol["M"].ik_rating == 8
    assert by_symbol["/E"].fixture_type == "emergency conversion kit"
    assert by_symbol["/E"].emergency_hours == 3


def test_fixture_symbol_does_not_become_wattage() -> None:
    assert WATT_RE.search("SL150W") is None
    assert WATT_RE.search("SL1 50W") is not None


def test_builds_excel_and_pdf_quotes() -> None:
    request = make_quote()

    xlsx = build_xlsx(request)
    workbook = load_workbook(BytesIO(xlsx))
    sheet = workbook["Product Quotation"]
    assert sheet["A1"].value == "TECS LIGHTING QUOTATION"
    assert sheet["B3"].value == "Test Project"
    assert sheet["D7"].value == "Sample downlight"
    assert sheet["F7"].value == 25

    pdf = build_pdf(request)
    reader = PdfReader(BytesIO(pdf))
    assert len(reader.pages) == 1
    assert "TECS LIGHTING QUOTATION" in (reader.pages[0].extract_text() or "")
    assert "Sample downlight" in (reader.pages[0].extract_text() or "")


def test_builds_individual_compliance_sheets() -> None:
    request = TechnicalSheetRequest(
        project=ProjectDetails(
            project_name="New Car Showroom",
            client="Example Client",
            consultant="Example Consultant",
            contractor="Example Contractor",
            reference="TC-001",
        ),
        items=[
            TechnicalItem(
                id="f1",
                fitting_type="F1, F1E",
                brand="Dialight",
                product_name="Hazardous linear fitting",
                country_of_origin="United Kingdom",
                model_no="HZ-100",
                rows=[
                    ComplianceRow(
                        parameter="IP Rating / IK Rating",
                        specified="IP66 / IK10",
                        proposed="IP66 / IK10",
                        status="complies",
                    ),
                    ComplianceRow(
                        parameter="Lamp / Lumen / Color Temp / Efficacy",
                        specified="8 W",
                        proposed="10 W",
                        status="deviation",
                        remarks="Offered input power is 2 W higher.",
                    ),
                ],
            )
        ],
    )

    xlsx = build_compliance_xlsx(request)
    workbook = load_workbook(BytesIO(xlsx))
    sheet = workbook["F1, F1E"]
    assert sheet["A8"].value == "TECHNICAL DATA SHEET"
    assert sheet["C25"].value == "IP66 / IK10"
    assert sheet["D25"].value == "IP66 / IK10"
    assert "DEVIATION" in sheet["E21"].value
    assert "COMPLIES - Complies" not in " ".join(
        str(sheet[f"E{row_number}"].value or "") for row_number in range(13, 29)
    )
    assert sheet.column_dimensions["B"].width == 1.55
    assert sheet.row_dimensions[13].height == 334.8
    assert len(sheet._images) == 1

    multi_request = request.model_copy(deep=True)
    multi_request.items.append(
        request.items[0].model_copy(
            update={
                "id": "f2",
                "fitting_type": "F2, F2E",
                "rows": [
                    ComplianceRow(
                        parameter="Description",
                        proposed="Proposed LED luminaire",
                        status="complies",
                    )
                ],
            },
            deep=True,
        )
    )
    multi_workbook = load_workbook(BytesIO(build_compliance_xlsx(multi_request)))
    assert multi_workbook.sheetnames == ["F1, F1E", "F2, F2E"]
    assert len(multi_workbook["F2, F2E"]._images) == 1
    assert multi_workbook["F2, F2E"]["D10"].value == "F2, F2E"
    assert multi_workbook["F2, F2E"]["D13"].value == "Proposed LED luminaire"
    assert multi_workbook["F2, F2E"]["C21"].value is None
    assert multi_workbook["F2, F2E"]["D21"].value is None
    assert multi_workbook["F2, F2E"]["C25"].value is None
    assert multi_workbook["F2, F2E"]["D25"].value is None

    pdf = build_compliance_pdf(request)
    reader = PdfReader(BytesIO(pdf))
    assert len(reader.pages) == 1
    text = reader.pages[0].extract_text() or ""
    assert "TECHNICAL DATA SHEET" in text
    assert "Offered input power is 2 W higher" in text


def test_builds_commercial_workbook_from_exact_template() -> None:
    request = CommercialQuotationRequest(
        project=ProjectDetails(
            project_name="New Car Showroom",
            client="Example Client",
            consultant="Example Consultant",
            contractor="Example Contractor",
            reference="CQ-001",
        ),
        currency="GBP",
        exchange_rates={"EUR": 0.85},
        items=[
            TechnicalItem(
                id="f1",
                fitting_type="F1",
                quantity=8,
                brand="LuxeLED",
                product_name="Mellow III Backlit Panel",
                country_of_origin="United Kingdom",
                model_no="MEL-III-4K-40W",
                unit_price=27.5,
                unit_price_currency="EUR",
                rows=[
                    ComplianceRow(
                        parameter="Description",
                        proposed="595 x 595 mm recessed LED panel",
                        status="complies",
                    )
                ],
            ),
            TechnicalItem(
                id="f2",
                fitting_type="F2",
                quantity=4,
                brand="Signify",
                product_name="Recessed downlight",
                model_no="DN100",
                unit_price=None,
            ),
        ],
    )

    workbook = load_workbook(BytesIO(build_commercial_xlsx(request)), data_only=False)
    assert workbook.sheetnames == ["Costing", "Offer"]
    for sheet in workbook.worksheets:
        assert sheet["C6"].value == "Example Contractor"
        assert sheet["C7"].value == "Consultant : Example Consultant"
        assert sheet["C9"].value == "PROJECT : New Car Showroom"
        assert sheet["C12"].value == "REFERENCE: CQ-001"
        assert sheet["C16"].value == "595 x 595 mm recessed LED panel"
        assert sheet["D16"].value == "LuxeLED - United Kingdom"
        assert sheet["H16"].value == "Euro"
        assert sheet["I16"].value == 27.5
        assert sheet["J16"].value == 1
        assert sheet["K16"].value == 0.85
        assert sheet["L16"].value == 1.15
        assert sheet["M16"].value == '=IF(I16="","",F16*I16*J16*K16*(L16-1))'
        assert sheet["N16"].value == 1.07
        assert sheet["O16"].value == 0
        assert sheet["P16"].value == '=IF(I16="","",(I16*J16*K16*L16*N16)+O16)'
        assert sheet["Q16"].value == '=IF(P16="","",F16*P16)'
        assert sheet["R16"].value == '=IF(P16="","",ROUNDUP(P16/0.7,0))'
        assert sheet["S16"].value == '=IF(R16="","",R16*F16)'
        assert sheet["I17"].value is None
        assert sheet["H17"].value is None
        assert sheet["K17"].value is None
        assert sheet["S17"].value == '=IF(R17="","",R17*F17)'
        assert sheet["R14"].value == "U.Price (GBP)"
        assert sheet["S14"].value == "T.Price\n(GBP)"
        assert sheet["S135"].value == '=IF(COUNT(I16:I17)=0,"",SUM(S16:S17))'
        assert sheet.row_dimensions[18].hidden is True
    assert workbook["Offer"]["B135"].value == "OFFER VALUE IN GBP"
    assert workbook["Costing"]["B135"].value is None
    assert workbook["Offer"].column_dimensions["H"].hidden is True


def test_commercial_export_saves_to_configured_download_folder(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setenv("TECS_EXPORT_DIR", str(tmp_path))
    request = CommercialQuotationRequest(
        project=ProjectDetails(project_name="Commercial Export Test"),
        currency="EUR",
        items=[
            TechnicalItem(
                id="f1",
                fitting_type="F1",
                quantity=2,
                brand="Whitecroft Lighting",
                product_name="Tegan 2",
                model_no="TG2-01",
            )
        ],
    )
    client = TestClient(app)

    response = client.post(
        "/api/commercial/xlsx/save?filename=TECS-Commercial-Quotation",
        json=request.model_dump(mode="json"),
    )

    assert response.status_code == 200
    saved_path = Path(response.json()["path"])
    assert saved_path == tmp_path / "TECS-Commercial-Quotation.xlsx"
    workbook = load_workbook(saved_path, data_only=False)
    assert workbook["Offer"]["R14"].value == "U.Price (EUR)"
    assert workbook["Offer"]["I16"].value is None


def test_commercial_export_requires_rate_for_priced_foreign_currency() -> None:
    request = CommercialQuotationRequest(
        project=ProjectDetails(project_name="Exchange Rate Test"),
        currency="OMR",
        exchange_rates={},
        items=[
            TechnicalItem(
                id="f1",
                fitting_type="F1",
                quantity=1,
                brand="Whitecroft Lighting",
                product_name="Tegan 2",
                unit_price=50,
                unit_price_currency="EUR",
            )
        ],
    )

    with pytest.raises(ValueError, match="EUR to OMR exchange rate"):
        build_commercial_xlsx(request)


def test_compliance_export_saves_to_configured_download_folder(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setenv("TECS_EXPORT_DIR", str(tmp_path))
    request = TechnicalSheetRequest(
        project=ProjectDetails(project_name="Windows Export Test"),
        items=[
            TechnicalItem(
                id="f1",
                fitting_type="F1",
                brand="LuxeLED",
                product_name="Mellow III",
                model_no="MEL III-4K-40W-595-IP54-S",
                rows=[
                    ComplianceRow(
                        parameter="Wattage",
                        specified="40 W",
                        proposed="40 W",
                        status="complies",
                    )
                ],
            )
        ],
    )
    client = TestClient(app)

    response = client.post(
        "/api/compliance/xlsx/save?filename=F1-Technical-Compliance",
        json=request.model_dump(mode="json"),
    )

    assert response.status_code == 200
    saved_path = Path(response.json()["path"])
    assert saved_path == tmp_path / "F1-Technical-Compliance.xlsx"
    assert saved_path.read_bytes().startswith(b"PK")


def test_export_folder_setting_can_be_read(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("TECS_EXPORT_DIR", str(tmp_path))
    client = TestClient(app)

    response = client.get("/api/settings/export-folder")

    assert response.status_code == 200
    assert response.json()["path"] == str(tmp_path)


def test_health_and_quote_endpoints(monkeypatch) -> None:
    monkeypatch.setattr("tecs_engine.main.has_api_key", lambda: False)
    client = TestClient(app)
    health = client.get("/api/health")
    assert health.status_code == 200
    assert health.json()["status"] == "ok"
    assert health.json()["api_key_configured"] is False
    assert health.json()["local_ai"]["state"] in {
        "ready", "starting", "not_installed", "error"
    }

    response = client.post("/api/quote/pdf", json=make_quote().model_dump(mode="json"))
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    assert response.content.startswith(b"%PDF")


def test_catalog_browse_endpoint_does_not_require_api_research(monkeypatch) -> None:
    client = TestClient(app)
    request = ProductSearchRequest(fixture=make_fixture(), brand="LuxeLED")

    response = client.post("/api/catalog/browse", json=request.model_dump(mode="json"))

    assert response.status_code == 200
    assert set(response.json()) == {
        "products", "families", "facets", "requirement_family", "freshness_days"
    }


def test_api_key_can_be_saved_and_removed(monkeypatch) -> None:
    test_api_key = "test-api-key-value-long-enough-for-validation"
    saved: list[str] = []
    monkeypatch.setattr("tecs_engine.main.save_api_key", saved.append)
    monkeypatch.setattr("tecs_engine.main.delete_api_key", lambda: False)
    client = TestClient(app)

    save_response = client.post(
        "/api/settings/api-key",
        json={"api_key": test_api_key},
    )
    assert save_response.status_code == 200
    assert saved == [test_api_key]

    remove_response = client.delete("/api/settings/api-key")
    assert remove_response.status_code == 200
    assert remove_response.json() == {
        "removed": True,
        "api_key_configured": False,
    }
