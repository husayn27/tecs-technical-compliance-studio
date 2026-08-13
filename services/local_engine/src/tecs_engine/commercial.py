from __future__ import annotations

from copy import copy
from importlib.resources import files
from io import BytesIO

from openpyxl import load_workbook

from .models import CommercialQuotationRequest, TechnicalItem

FIRST_ITEM_ROW = 16
LAST_ITEM_ROW = 134
TOTAL_ROW = 135
VAT_ROW = 136
GRAND_TOTAL_ROW = 137


def _display_currency(currency: str) -> str:
    return "Euro" if currency == "EUR" else currency


def _description(item: TechnicalItem) -> str:
    proposed = next(
        (
            row.proposed.strip()
            for row in item.rows
            if row.parameter == "Description" and row.proposed.strip()
        ),
        "",
    )
    return proposed or item.product_name or item.fitting_type


def _make_and_origin(item: TechnicalItem) -> str:
    values = [value.strip() for value in (item.brand, item.country_of_origin) if value.strip()]
    return " - ".join(values)


def _clear_item_area(sheet) -> None:
    for row in range(FIRST_ITEM_ROW, LAST_ITEM_ROW + 1):
        sheet.row_dimensions[row].hidden = False
        for column in range(1, 20):
            sheet.cell(row=row, column=column).value = None


def _copy_item_style(sheet, source_row: int, target_row: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, 20):
        source = sheet.cell(row=source_row, column=column)
        target = sheet.cell(row=target_row, column=column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def _populate_sheet(sheet, request: CommercialQuotationRequest) -> None:
    items = [item for item in request.items if item.selected]
    if not items:
        raise ValueError("Select at least one product for the commercial quotation.")
    if len(items) > LAST_ITEM_ROW - FIRST_ITEM_ROW + 1:
        raise ValueError("The commercial quotation template supports up to 119 products.")

    offer_currency = _display_currency(request.currency)
    sheet["C6"] = request.project.contractor or ""
    sheet["C7"] = (
        f"Consultant : {request.project.consultant}"
        if request.project.consultant
        else ""
    )
    sheet["C9"] = f"PROJECT : {request.project.project_name}" if request.project.project_name else "PROJECT :"
    sheet["C10"] = f"Client : {request.project.client}" if request.project.client else "Client :"
    sheet["C12"] = f"REFERENCE: {request.project.reference}" if request.project.reference else "REFERENCE:"
    sheet["C15"] = ""
    sheet["R14"] = f"U.Price ({request.currency})"
    sheet["S14"] = f"T.Price\n({request.currency})"

    _clear_item_area(sheet)
    for offset, item in enumerate(items):
        row = FIRST_ITEM_ROW + offset
        if row != FIRST_ITEM_ROW:
            _copy_item_style(sheet, FIRST_ITEM_ROW, row)
        sheet.cell(row=row, column=1, value=offset + 1)
        sheet.cell(row=row, column=2, value=item.fitting_type)
        sheet.cell(row=row, column=3, value=_description(item))
        sheet.cell(row=row, column=4, value=_make_and_origin(item))
        sheet.cell(row=row, column=5, value=item.model_no)
        sheet.cell(row=row, column=6, value=item.quantity)
        sheet.cell(row=row, column=7, value="Nos")
        unit_currency_code = item.unit_price_currency or request.currency
        unit_currency = _display_currency(unit_currency_code)
        exchange_rate = (
            1.0
            if unit_currency_code == request.currency
            else request.exchange_rates.get(unit_currency_code)
        )
        if item.unit_price is not None and exchange_rate is None:
            raise ValueError(
                f"Enter the {unit_currency_code} to {request.currency} exchange rate "
                f"for {item.fitting_type or item.product_name}."
            )
        if exchange_rate is not None and exchange_rate <= 0:
            raise ValueError(
                f"The {unit_currency_code} to {request.currency} exchange rate must be greater than zero."
            )
        sheet.cell(row=row, column=8, value=unit_currency if item.unit_price is not None else None)
        sheet.cell(row=row, column=9, value=item.unit_price)
        sheet.cell(row=row, column=10, value=1)
        sheet.cell(row=row, column=11, value=exchange_rate if item.unit_price is not None else None)
        sheet.cell(row=row, column=12, value=1.15)
        sheet.cell(row=row, column=13, value=f'=IF(I{row}="","",F{row}*I{row}*J{row}*K{row}*(L{row}-1))')
        sheet.cell(row=row, column=14, value=1.07)
        sheet.cell(row=row, column=15, value=0)
        sheet.cell(row=row, column=16, value=f'=IF(I{row}="","",(I{row}*J{row}*K{row}*L{row}*N{row})+O{row})')
        sheet.cell(row=row, column=17, value=f'=IF(P{row}="","",F{row}*P{row})')
        sheet.cell(row=row, column=18, value=f'=IF(P{row}="","",ROUNDUP(P{row}/0.7,0))')
        sheet.cell(row=row, column=19, value=f'=IF(R{row}="","",R{row}*F{row})')

    first_unused = FIRST_ITEM_ROW + len(items)
    for row in range(first_unused, LAST_ITEM_ROW + 1):
        sheet.row_dimensions[row].hidden = True

    last_item = FIRST_ITEM_ROW + len(items) - 1
    sheet[f"F{TOTAL_ROW}"] = f"=SUM(F{FIRST_ITEM_ROW}:F{last_item})"
    sheet[f"M{TOTAL_ROW}"] = f'=IF(COUNT(I{FIRST_ITEM_ROW}:I{last_item})=0,"",SUM(M{FIRST_ITEM_ROW}:M{last_item}))'
    sheet[f"P{TOTAL_ROW}"] = "Cost"
    sheet[f"Q{TOTAL_ROW}"] = f'=IF(COUNT(I{FIRST_ITEM_ROW}:I{last_item})=0,"",SUM(Q{FIRST_ITEM_ROW}:Q{last_item}))'
    sheet[f"P{VAT_ROW}"] = "G.P Value"
    sheet[f"Q{VAT_ROW}"] = f'=IF(S{TOTAL_ROW}="","",S{TOTAL_ROW}-Q{TOTAL_ROW})'
    sheet[f"P{GRAND_TOTAL_ROW}"] = "G.P%"
    sheet[f"Q{GRAND_TOTAL_ROW}"] = f'=IF(S{TOTAL_ROW}="","",Q{VAT_ROW}/S{TOTAL_ROW})'
    sheet[f"S{TOTAL_ROW}"] = f'=IF(COUNT(I{FIRST_ITEM_ROW}:I{last_item})=0,"",SUM(S{FIRST_ITEM_ROW}:S{last_item}))'
    sheet[f"S{VAT_ROW}"] = f'=IF(S{TOTAL_ROW}="","",S{TOTAL_ROW}*5%)'
    sheet[f"S{GRAND_TOTAL_ROW}"] = f'=IF(S{TOTAL_ROW}="","",S{TOTAL_ROW}+S{VAT_ROW})'
    if sheet.title == "Offer":
        sheet[f"B{TOTAL_ROW}"] = f"OFFER VALUE IN {offer_currency.upper()}"
        sheet[f"B{VAT_ROW}"] = f"VAT 5% IN {offer_currency.upper()}"
        sheet[f"B{GRAND_TOTAL_ROW}"] = f"TOTAL OFFER VALUE IN {offer_currency.upper()}"
    sheet.print_area = f"A5:S152"


def build_commercial_xlsx(request: CommercialQuotationRequest) -> bytes:
    template = files("tecs_engine").joinpath("assets/commercial-offer-template.xlsx")
    workbook = load_workbook(template)
    for sheet in workbook.worksheets:
        _populate_sheet(sheet, request)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
