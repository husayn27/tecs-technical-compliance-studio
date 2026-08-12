from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import QuoteRequest

NAVY = "17324D"
TEAL = "00A7A0"


def build_xlsx(request: QuoteRequest) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Product Quotation"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:G1")
    sheet["A1"] = "TECS LIGHTING QUOTATION"
    sheet["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet["A3"] = "Project"
    sheet["B3"] = request.project_name
    sheet["A4"] = "Customer"
    sheet["B4"] = request.customer_name or "-"
    sheet["E3"] = "Reference"
    sheet["F3"] = request.reference or "-"

    headers = ["Item", "Symbol", "Brand", "Selected product", "Product code", "Quantity", "Product link"]
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row=6, column=column, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(vertical="center")

    for index, line in enumerate(request.lines, start=1):
        product = line.product
        fixture = line.fixture
        values = [
            index,
            fixture.symbol,
            product.brand,
            product.product_name,
            product.product_code or "-",
            fixture.quantity,
            str(product.product_url),
        ]
        row = 6 + index
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column == 7:
                cell.hyperlink = value
                cell.style = "Hyperlink"

    widths = [8, 12, 18, 34, 22, 12, 48]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A7"
    bottom = 6 + len(request.lines)
    thin = Side(style="thin", color="D9E1E8")
    for row in sheet.iter_rows(min_row=6, max_row=bottom, min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(bottom=thin)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_pdf(request: QuoteRequest) -> bytes:
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="TECS Lighting Quotation",
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "TitleTECS", parent=styles["Title"], fontName="Helvetica-Bold",
        fontSize=20, leading=24, textColor=colors.HexColor(f"#{NAVY}"), alignment=TA_LEFT,
    )
    small = ParagraphStyle("Small", parent=styles["BodyText"], fontSize=8, leading=10)
    story = [
        Paragraph("TECS LIGHTING QUOTATION", title),
        Spacer(1, 4 * mm),
        Paragraph(f"<b>Project:</b> {request.project_name}", styles["BodyText"]),
        Paragraph(f"<b>Customer:</b> {request.customer_name or '-'} &nbsp;&nbsp; <b>Reference:</b> {request.reference or '-'}", styles["BodyText"]),
        Spacer(1, 5 * mm),
    ]
    data = [["Item", "Symbol", "Brand", "Selected product", "Product code", "Qty", "Product link"]]
    for index, line in enumerate(request.lines, start=1):
        data.append([
            str(index),
            line.fixture.symbol,
            line.product.brand,
            Paragraph(line.product.product_name, small),
            line.product.product_code or "-",
            str(line.fixture.quantity),
            Paragraph(f'<link href="{line.product.product_url}" color="#007A75">Open product</link>', small),
        ])
    table = Table(data, colWidths=[13 * mm, 18 * mm, 28 * mm, 65 * mm, 36 * mm, 14 * mm, 38 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{TEAL}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F7F9")]),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9E1E8")),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    document.build(story)
    return output.getvalue()

