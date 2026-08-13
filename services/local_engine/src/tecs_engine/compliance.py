from __future__ import annotations

from html import escape
from io import BytesIO
from importlib.resources import as_file, files
from copy import copy, deepcopy
import re

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import ComplianceRow, TechnicalItem, TechnicalSheetRequest

TECS_RED = "E71938"
LIGHT_GREY = "D9D9D9"
STATUS_FILL = {
    "complies": "E2F3EA",
    "deviation": "FDE7E7",
    "pending": "FFF2CC",
    "not_applicable": "EDEDED",
}
STATUS_LABEL = {
    "complies": "COMPLIES",
    "deviation": "DEVIATION",
    "pending": "PENDING REVIEW",
    "not_applicable": "NOT APPLICABLE",
}


def _safe_sheet_name(value: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "-", value).strip() or "Item"
    base = base[:31]
    candidate = base
    counter = 2
    while candidate in used:
        suffix = f"-{counter}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def _remark(row: ComplianceRow) -> str:
    if row.remarks.strip():
        return row.remarks.strip()
    if row.status == "complies":
        return "Complies."
    if row.status == "deviation":
        return "Deviation; refer to the proposed value."
    if row.status == "not_applicable":
        return "Not applicable."
    return "Engineer to confirm."


def _status_text(row: ComplianceRow) -> str:
    remark = _remark(row)
    boilerplate = {
        "complies",
        "not applicable",
        "engineer to confirm",
        "not published; engineer to confirm",
        "deviation; refer to the proposed value",
    }
    if remark.rstrip(".").strip().lower() in boilerplate:
        return STATUS_LABEL[row.status]
    return f"{STATUS_LABEL[row.status]} - {remark}"


def _populate_sheet(sheet, request: TechnicalSheetRequest, item: TechnicalItem) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.freeze_panes = "A12"

    widths = [24, 58, 58, 32]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.merge_cells("A1:D2")
    sheet["A1"] = "TECS   TECHNICAL SUPPLIES"
    sheet["A1"].font = Font(name="Arial", size=20, bold=True, color=TECS_RED)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

    sheet.merge_cells("A3:D3")
    sheet["A3"] = "PROJECT DETAILS"
    sheet["A3"].font = Font(name="Arial", size=13, bold=True)
    sheet["A3"].fill = PatternFill("solid", fgColor=LIGHT_GREY)
    sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")

    project_rows = [
        ("PROJECT NAME", request.project.project_name),
        ("CLIENT", request.project.client or "-"),
        ("CONSULTANT", request.project.consultant or "-"),
        ("CONTRACTOR", request.project.contractor or "-"),
    ]
    for row_index, (label, value) in enumerate(project_rows, start=4):
        sheet[f"A{row_index}"] = label
        sheet[f"A{row_index}"].font = Font(name="Arial", size=10, bold=True)
        sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=4)
        sheet.cell(row=row_index, column=2, value=value)
        sheet.cell(row=row_index, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[row_index].height = 28

    sheet.merge_cells("A8:D8")
    sheet["A8"] = "TECHNICAL DATA SHEET"
    sheet["A8"].font = Font(name="Arial", size=13, bold=True)
    sheet["A8"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells("A9:D9")
    sheet["A9"] = f"Fitting Type: {item.fitting_type}    |    Brand: {item.brand or '-'}    |    Model: {item.model_no or '-'}"
    sheet["A9"].font = Font(name="Arial", size=11, bold=True)
    sheet["A9"].fill = PatternFill("solid", fgColor=LIGHT_GREY)
    sheet["A9"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["PARAMETER", "SPECIFIED", "PROPOSED", "REMARKS"]
    for column, label in enumerate(headers, start=1):
        cell = sheet.cell(row=11, column=column, value=label)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="4F4F4F")
    for index, row in enumerate(item.rows, start=12):
        values = [row.parameter, row.specified or "-", row.proposed or "-", _status_text(row)]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=index, column=column, value=value)
            cell.font = Font(name="Arial", size=9, bold=column == 1)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        sheet.cell(row=index, column=4).fill = PatternFill("solid", fgColor=STATUS_FILL[row.status])
        longest = max(len(str(value)) for value in values)
        sheet.row_dimensions[index].height = min(120, max(28, 15 + (longest // 70) * 13))

    for row in sheet.iter_rows(min_row=3, max_row=11 + len(item.rows), min_col=1, max_col=4):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.print_area = f"A1:D{11 + len(item.rows)}"
    sheet.oddFooter.center.text = f"TECS Technical Supplies | {request.project.reference or request.project.project_name}"
    sheet.oddFooter.center.size = 8


def build_compliance_xlsx(request: TechnicalSheetRequest) -> bytes:
    items = [item for item in request.items if item.selected]
    if not items:
        raise ValueError("Select at least one product before exporting.")

    template_resource = files("tecs_engine").joinpath(
        "assets/technical-compliance-template.xlsx"
    )
    with as_file(template_resource) as template_path:
        workbook = load_workbook(template_path)

    base = workbook["F1, F1E"]
    template_images = [(image._data(), deepcopy(image.anchor)) for image in base._images]
    base._images = []
    for sheet in list(workbook.worksheets):
        if sheet is not base:
            workbook.remove(sheet)

    # Create every item sheet from the untouched template before populating any
    # product. Otherwise later copies inherit values written for the first item.
    item_sheets = [base]
    for _ in items[1:]:
        copied = workbook.copy_worksheet(base)
        copied.page_margins = copy(base.page_margins)
        copied.page_setup = copy(base.page_setup)
        copied.print_options = copy(base.print_options)
        copied.sheet_properties = copy(base.sheet_properties)
        copied.sheet_format = copy(base.sheet_format)
        copied.freeze_panes = base.freeze_panes
        copied.print_area = base.print_area
        item_sheets.append(copied)

    used: set[str] = set()
    for sheet, item in zip(item_sheets, items, strict=True):
        sheet._images = []
        for image_data, anchor in template_images:
            duplicated = XLImage(BytesIO(image_data))
            duplicated.anchor = deepcopy(anchor)
            sheet.add_image(duplicated)
        sheet.title = _safe_sheet_name(item.fitting_type, used)
        _populate_template_sheet(sheet, request, item)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _populate_template_sheet(sheet, request: TechnicalSheetRequest, item: TechnicalItem) -> None:
    sheet["B3"] = request.project.project_name
    sheet["B4"] = request.project.client or "-"
    sheet["B5"] = request.project.consultant or "-"
    sheet["B6"] = request.project.contractor or "-"
    sheet["D10"] = item.fitting_type

    row_map = {
        "Description": 13,
        "Make": 14,
        "Country of Origin": 15,
        "Model No": 16,
        "Mounting": 17,
        "Housing / Construction": 18,
        "Reflector / Optical System": 19,
        "Control Gear / Ballast": 20,
        "Lamp / Lumen / Color Temp / Efficacy": 21,
        "Emergency": 22,
        "CRI": 23,
        "LED life": 24,
        "IP Rating / IK Rating": 25,
        "UGR": 26,
        "Finish": 27,
        "Remarks": 28,
    }
    # The source template contains example VLOOKUP formulas and cached values.
    # Exports are driven only by the current project, so optional fields that
    # were not supplied must remain blank instead of leaking template examples.
    for row_number in row_map.values():
        for column in ("C", "D", "E"):
            cell = sheet[f"{column}{row_number}"]
            if not isinstance(cell, MergedCell):
                cell.value = None
    by_parameter = {row.parameter: row for row in item.rows}
    for parameter, row_number in row_map.items():
        row = by_parameter.get(parameter)
        if row is None:
            continue
        if parameter == "Remarks":
            combined = "\n".join(value for value in (row.specified, row.proposed) if value.strip())
            sheet[f"C{row_number}"] = combined or "-"
        else:
            sheet[f"C{row_number}"] = row.specified or "-"
            sheet[f"D{row_number}"] = row.proposed or "-"
        sheet[f"E{row_number}"] = _status_text(row)

    if item.brand:
        sheet["D14"] = item.brand
    if item.country_of_origin:
        sheet["D15"] = item.country_of_origin
    if item.model_no:
        sheet["D16"] = item.model_no

    location = by_parameter.get("Location")
    if location and any((location.specified.strip(), location.proposed.strip(), location.remarks.strip())):
        location_text = "Location: " + " / ".join(
            value for value in (location.specified, location.proposed) if value.strip()
        )
        existing = sheet["C28"].value
        sheet["C28"] = f"{existing}\n{location_text}" if existing and existing != "-" else location_text


def _p(value: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape(value or "-").replace("\n", "<br/>"), style)


def build_compliance_pdf(request: TechnicalSheetRequest) -> bytes:
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=7 * mm,
        bottomMargin=7 * mm,
        title="TECS Technical Compliance Sheets",
    )
    styles = getSampleStyleSheet()
    body = ParagraphStyle("ComplianceBody", parent=styles["BodyText"], fontName="Helvetica", fontSize=5.6, leading=6.4, alignment=TA_LEFT)
    label = ParagraphStyle("ComplianceLabel", parent=body, fontName="Helvetica-Bold")
    center = ParagraphStyle("ComplianceCenter", parent=body, fontName="Helvetica-Bold", alignment=TA_CENTER, fontSize=7.2, leading=8)
    title = ParagraphStyle("TECS", parent=center, textColor=colors.HexColor(f"#{TECS_RED}"), fontSize=15, leading=16, alignment=TA_LEFT)
    story = []
    items = [item for item in request.items if item.selected]
    if not items:
        raise ValueError("Select at least one product before exporting.")
    for item_index, item in enumerate(items):
        if item_index:
            story.append(PageBreak())
        story.extend([
            Paragraph("TECS   TECHNICAL SUPPLIES", title),
            Spacer(1, 0.5 * mm),
            Table([[Paragraph("PROJECT DETAILS", center)]], colWidths=[281 * mm]),
        ])
        project_data = [
            [_p("PROJECT NAME", label), _p(request.project.project_name, center)],
            [_p("CLIENT", label), _p(request.project.client or "-", center)],
            [_p("CONSULTANT", label), _p(request.project.consultant or "-", center)],
            [_p("CONTRACTOR", label), _p(request.project.contractor or "-", center)],
        ]
        project_table = Table(project_data, colWidths=[45 * mm, 236 * mm], rowHeights=[6.5 * mm] * 4)
        project_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#555555")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ]))
        story.extend([project_table, Spacer(1, 0.7 * mm), Paragraph("TECHNICAL DATA SHEET", center)])
        fitting = Table([[_p(f"Fitting Type: {item.fitting_type}", center), _p(f"{item.brand or '-'} | {item.model_no or '-'}", center)]], colWidths=[90 * mm, 191 * mm])
        fitting.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{LIGHT_GREY}")), ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#555555")), ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        story.append(fitting)

        data = [[_p("PARAMETER", center), _p("SPECIFIED", center), _p("PROPOSED", center), _p("REMARKS", center)]]
        for row in item.rows:
            data.append([
                _p(row.parameter, label),
                _p(row.specified or "-", body),
                _p(row.proposed or "-", body),
                _p(_status_text(row), body),
            ])
        table = Table(data, colWidths=[39 * mm, 88 * mm, 88 * mm, 66 * mm], repeatRows=1)
        commands = [
            ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#555555")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.white),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 1.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ]
        for row_index, row in enumerate(item.rows, start=1):
            commands.append(("BACKGROUND", (3, row_index), (3, row_index), colors.HexColor(f"#{STATUS_FILL[row.status]}")))
        table.setStyle(TableStyle(commands))
        story.append(table)
    document.build(story)
    return output.getvalue()
